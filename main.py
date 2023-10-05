import logging
import time
from io import StringIO
import pandas as pd
from openpyxl import load_workbook
from selenium_stealth import stealth
import random
from datetime import datetime
import requests
import re
from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class FSSPScraper:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)
        self.logger.addHandler(logging.FileHandler('scraper.log'))
        self.logger.info("FSSPScraper initialized")

        self.TASK_PATH_EXCEL = 'names.xlsx'
        self.TIMEOUT = 5
        self.URL = 'http://fssprus.ru/'
        self.MODAL_CLOSE = '//*[contains(@class, "ModalClose")]'
        self.ADVANCED_SEARCH_BUTTON = '//*[contains(@class, "Button")]//*[contains(text(), "Расширенный поиск")]'
        self.INDIVIDUAL_SEARCH = '//*[contains(@class, "Button")]//*[contains(text(), "Поиск физич")]'
        self.FIELD_LASTNAME = '//*[contains(@placeholder, "Введите фамилию")]'
        self.FIELD_NAME = '//*[contains(@placeholder, "Введите имя")]'
        self.FIELD_MIDDLENAME = '//*[contains(@placeholder, "Введите отчество")]'
        self.FIELD_BIRTHDAY = '//*[contains(@placeholder, "Выберите дату")]'
        self.FIELD_REGIONS_SELECT = '//*[contains(@placeholder, "Выберите регион")]'
        self.REGION = 'Все регионы'
        self.SEARCH_BUTTON = '//*[contains(@href, "region")]//*[contains(text(), "Найти")]'
        self.CAPTCHA_IMG = '//*[@id="capchaVisual"]'
        self.CAPTCHA_INPUT = '//*[contains(@id, "code")]'
        self.CAPTCHA_ERROR = '//*[contains(@class, "error")]/following-sibling::*[contains(text(), "Неверно")]'
        self.CAPTCHA_SUBMIT = '//*[contains(@id, "submit")]'
        self.URL_SOLVE_CAPTCHA = "http://iamnotbot.com:5000/createTask"
        self.RESULTS_WARNING = '//*[contains(@class, "results")]//*[contains(@class, "warning")]'
        self.RESULTS_EMPTY = '//*[contains(@class, "results")]//*[contains(@class, "empty")]'
        self.RESULTS_NOT_FOUND = '//*[contains(@class, "results")]//*[contains(text(), "не найдено")]'
        self.RESULTS_TABLE = '//*[@id="content"]//table'

        user_agents = [
            # Вставить ваш список пользовательских агентов здесь
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
            'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.1 Safari/605.1.15',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 13_1) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.1 Safari/605.1.15',
        ]
        self.user_agent = random.choice(user_agents)

        self.driver = None

    def setup_driver(self):
        # Создать объект опций для браузера Chrome и настроить его для успешной работы с сайтом
        chrome_options = ChromeOptions()
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_argument('--disable-popup-blocking')
        chrome_options.add_argument('--start-maximized')
        chrome_options.add_argument('--disable-extensions')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')

        self.driver = webdriver.Chrome(options=chrome_options)
        self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        chrome_options.add_argument(f'user-agent={self.user_agent}')
        self.driver.execute_cdp_cmd('Network.setUserAgentOverride', {"userAgent": self.user_agent})

    def search_individual(self, last_name, first_name, middle_name, birth_date):
        if not self.driver:
            self.setup_driver()

        wait = WebDriverWait(self.driver, self.TIMEOUT)
        self.driver.get(self.URL)

        try:
            wait.until(EC.presence_of_element_located((By.XPATH, self.MODAL_CLOSE))).click()
        except Exception as e:
            self.logger.error("Элемент не найден: " + self.MODAL_CLOSE + "\n--->\n" + str(e) + "<---")

        try:
            wait.until(EC.presence_of_element_located((By.XPATH, self.ADVANCED_SEARCH_BUTTON))).click()
        except Exception as e:
            self.logger.error("Не удалось выполнить клик на кнопке Расширенный поиск: " + self.ADVANCED_SEARCH_BUTTON)
            return

        try:
            wait.until(EC.presence_of_element_located((By.XPATH, self.INDIVIDUAL_SEARCH))).click()
        except Exception as e:
            self.logger.error("Не удалось выполнить клик на вкладке Поиск физического лица: " + self.INDIVIDUAL_SEARCH)
            return

        try:
            wait.until(EC.presence_of_element_located((By.XPATH, self.FIELD_LASTNAME))).send_keys(last_name)
            wait.until(EC.presence_of_element_located((By.XPATH, self.FIELD_NAME))).send_keys(first_name)
            wait.until(EC.presence_of_element_located((By.XPATH, self.FIELD_MIDDLENAME))).send_keys(middle_name)
            wait.until(EC.presence_of_element_located((By.XPATH, self.FIELD_BIRTHDAY))).send_keys(birth_date.strftime("%d.%m.%Y"))
            wait.until(EC.presence_of_element_located((By.XPATH, self.FIELD_REGIONS_SELECT))).send_keys(self.REGION + '\uE007')
            wait.until(EC.presence_of_element_located((By.XPATH, self.SEARCH_BUTTON))).click()
        except Exception as e:
            self.logger.error("Ошибка при заполнении форм поиска: " + str(e))
            return

        original_window = self.driver.current_window_handle
        try:
            wait.until(EC.number_of_windows_to_be(2))
        except Exception as e:
            self.logger.error("Ошибка при ожидании открытия второго окна или вкладки: " + str(e))
            return

        for window_handle in self.driver.window_handles:
            if window_handle != original_window:
                self.driver.switch_to.window(window_handle)
                break

        try:
            wait.until(EC.url_contains(birth_date.strftime("%d.%m.%Y")))
        except Exception as e:
            self.logger.error("Ошибка при ожидании загрузки страницы: " + str(e))
            return

        captcha_error = True
        while captcha_error:
            captcha_image_container = None
            try:
                captcha_image_container = wait.until(EC.presence_of_element_located((By.XPATH, self.CAPTCHA_IMG)))
            except Exception as e:
                self.logger.error("Ошибка при поиске CAPTCHA-изображения: " + str(e))
                return

            captcha_image_base64 = re.sub('^data:image/jpeg;base64,', '', captcha_image_container.get_attribute("src"))
            json = {"clientKey": "DEMO",
                    "task": {
                        "type": "ImageToTextTask",
                        "subType": "fssp",
                        "body": captcha_image_base64
                    }}
            request = requests.post(self.URL_SOLVE_CAPTCHA, json=json)
            if request.status_code == 200:
                captcha_solved = request.text
                self.logger.info("CAPTCHA = " + captcha_solved)

                try:
                    wait.until(EC.presence_of_element_located((By.XPATH, self.CAPTCHA_INPUT))).send_keys(captcha_solved)
                except Exception as e:
                    self.logger.error("Ошибка при вводе CAPTCHA: " + str(e))
                    return

                time.sleep(3)
                try:
                    wait.until(EC.presence_of_element_located((By.XPATH, self.CAPTCHA_SUBMIT))).click()
                except Exception as e:
                    self.logger.error("Ошибка при подтверждении CAPTCHA: " + str(e))
                    return

                try:
                    wait.until(EC.presence_of_element_located((By.XPATH, self.CAPTCHA_ERROR)))
                    continue
                except Exception as e:
                    self.logger.info("CAPTCHA успешно решена.")
                    captcha_error = False

        results_table = None
        if not captcha_error:
            try:
                wait.until(EC.presence_of_element_located((By.XPATH, self.RESULTS_WARNING))).is_displayed()
                wait.until(EC.presence_of_element_located((By.XPATH, self.RESULTS_EMPTY))).is_displayed()
                self.logger.info("Результаты поиска в ожидании или произошла ошибка. Запрос: "
                      + ' '.join([last_name, first_name, middle_name, birth_date.strftime('%d.%m.%Y')]))
            except Exception as e:
                try:
                    wait.until(EC.presence_of_element_located((By.XPATH, self.RESULTS_NOT_FOUND))).is_displayed()
                    self.logger.info("По данному запросу ничего не найдено: "
                          + ' '.join([last_name, first_name, middle_name, birth_date.strftime('%d.%m.%Y')]))
                except Exception as e:
                    try:
                        results_table = wait.until(EC.presence_of_element_located((By.XPATH, self.RESULTS_TABLE)))
                        self.logger.info("Найдены исполнительные производства в отношении: "
                              + ' '.join([last_name, first_name, middle_name, birth_date.strftime('%d.%m.%Y')]))
                    except Exception as e:
                        self.logger.error("Произошла ошибка поиска. Перехожу к следующему запросу.")
                        self.driver.close()
                        self.driver.switch_to.window(original_window)
                        return

        if results_table:
            table_html = results_table.get_attribute("outerHTML")
            try:
                df = pd.read_html(StringIO(table_html))[0]
                file_path = ' '.join([last_name, first_name, middle_name, birth_date.strftime('%d.%m.%Y')]) \
                            + "-" + datetime.now().strftime('%Y-%m-%d %H.%M.%S') + ".xlsx"
                try:
                    df.to_excel(file_path, index=False)
                    self.logger.info("Данные успешно сохранены в файле Excel: " + file_path)
                except Exception as e:
                    self.logger.error("Ошибка при сохранении данных в Excel: " + str(e))
            except Exception as e:
                self.logger.error("Ошибка при чтении таблицы: " + str(e))

        self.driver.close()
        self.driver.switch_to.window(original_window)
        self.logger.info("Запрос успешно обработан")

    def close_driver(self):
        if self.driver:
            self.driver.quit()

if __name__ == '__main__':
    scraper = FSSPScraper()
    wb = load_workbook(scraper.TASK_PATH_EXCEL)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        last_name, first_name, middle_name, birth_date = row
        scraper.search_individual(last_name, first_name, middle_name, birth_date)

    scraper.close_driver()
